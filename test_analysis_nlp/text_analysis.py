import requests
import pandas as pd
import numpy as np
import openpyxl
from bs4 import BeautifulSoup
import nltk
import os
import re
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords
nltk.download('stopwords')


def calculate_score(url):
    # Make a request to the website
    response = requests.get(url)


    if response.status_code == 200:
        # Create a Beautiful Soup object
        soup = BeautifulSoup(response.text, 'html.parser')

        # Find specific elements in the HTML
        title_2 = soup.find('h1').getText()
        paragraphs = soup.find_all('p')
        paragraphs = paragraphs[16:]

        # Load stop words
        stop_words = set(stopwords.words('english'))

        # Cleaning using Stop Words Lists
        # filter_paragraph = []
        # stopwords_directory = './StopWords'
        # for filename in os.listdir(stopwords_directory):
        #     if os.path.isfile(os.path.join(stopwords_directory, filename)):
        #         with open(os.path.join(stopwords_directory, filename), 'r', encoding='latin-1') as file:
        #             stop_words.update(file.read().splitlines())
        directory = './StopWords'  # Replace with your directory path
        filter_paragraph = []

        # Loop over file names in the directory
        for filename in os.listdir(directory):
            if os.path.isfile(os.path.join(directory, filename)):
                # Perform operations on the file
                encodings = ['utf-8', 'latin-1']  # Add more encodings if needed
                for encoding in encodings:
                    try:
                        with open(os.path.join(directory, filename), 'r', encoding=encoding) as file:
                            text = file.read()
                            for p in paragraphs:
                                filter_word = [str(word) for word in p.get_text().split() if word.lower() not in text]
                                filter_paragraph.append(' '.join(filter_word))
                        break  # Break out of the loop if the file is read successfully
                    except UnicodeDecodeError:
                        print(f"Unable to read {filename} with {encoding} encoding.")

        # Loop over paragraphs

        # Calculate metrics
        no_of_words = sum(len(p.split()) for p in filter_paragraph)
        sentence_lengths = [len(sent.split()) for p in filter_paragraph for sent in sent_tokenize(p) if sent]
        avg_sentence_length = sum(sentence_lengths) / len(sentence_lengths) if sentence_lengths else 0
        percentage_complex_words = (sum(1 for word in ' '.join(filter_paragraph).split() if count_syllables(word) > 2) / no_of_words) * 100 if no_of_words > 0 else 0
        fog_index = 0.4 * (avg_sentence_length + percentage_complex_words)
        avg_words_per_sentence = no_of_words / len(filter_paragraph) if filter_paragraph else 0
        complex_word_count = sum(1 for word in ' '.join(filter_paragraph).split() if count_syllables(word) > 2)
        syllables = sum(count_syllables(word) for word in ' '.join(filter_paragraph).split())
        syllables_per_word = syllables / no_of_words if no_of_words > 0 else 0

        # Personal Pronouns
        pronouns = ["I", "we", "my", "ours", "us"]
        personal_pronouns_count = sum(len(re.findall(r"\b{}\b".format(pronoun), ' '.join(filter_paragraph), re.IGNORECASE)) for pronoun in pronouns)

        # Average Word Length
        avg_word_length = syllables / no_of_words if no_of_words > 0 else 0

        # Create dictionaries for positive and negative words
        positive_words = set()
        negative_words = set()

        # Creating a dictionary of Positive and Negative words
        master_dictionary_directory = './MasterDictionary'
        for filename in os.listdir(master_dictionary_directory):
            if os.path.isfile(os.path.join(master_dictionary_directory, filename)):
                with open(os.path.join(master_dictionary_directory, filename), 'r', encoding='latin-1') as file:
                    words = file.read().splitlines()
                    if filename == 'positive-words.txt':
                        positive_words.update(words)
                    else:
                        negative_words.update(words)

        # Extracting Derived variables
        positive_score = sum(1 for word in ' '.join(filter_paragraph).split() if word in positive_words)
        negative_score = sum(-1 for word in ' '.join(filter_paragraph).split() if word in negative_words)
        negative_score *= -1
        polarity_score = (positive_score - negative_score) / (positive_score + negative_score + 0.000001)
        subjectivity_score = (positive_score + negative_score) / (no_of_words + 0.000001)
        # print("Title_2:", title_2)
        # print("Filtered Paragraphs:")
        # for p in filter_paragraph:
        #     print("-", p)
        
        # print("No. of Words:", no_of_words)
        # print("Average Sentence Length:", avg_sentence_length)
        # print("Percentage of Complex Words:", percentage_complex_words)
        # print("FOG Index:", fog_index)
        # print("Average Number of Words per Sentence:", avg_words_per_sentence)
        # print("Complex Word Count:", complex_word_count)
        # print("Syllables per Word:", syllables_per_word)
        # print("Personal Pronouns Count:", personal_pronouns_count)
        # print("Average Word Length:", avg_word_length)
        # print("Positive score:", positive_score)
        # print("Negative score:", negative_score)
        
        filter_paragraph_2=[]

        for p in filter_paragraph:
            words = word_tokenize(p)
            filter_word = [word for word in words if word.lower() not in stop_words]
            filter_paragraph_2.append(' '.join(filter_word))
        clened_words = sum(len(p.split()) for p in filter_paragraph_2)
        # print("Filtered Paragraphs(2):")
        # for p in filter_paragraph_2:
        #     print("-", p)
        # Print results
        # print("clened words",clened_words)
        data =pd.DataFrame( {
             "URL":[url],
            "No. of Words": [no_of_words],
            "Average Sentence Length": [avg_sentence_length],
            "Percentage of Complex Words": [percentage_complex_words],
            "FOG Index": [fog_index],
            "Average Number of Words per Sentence": [avg_words_per_sentence],
            "Complex Word Count": [complex_word_count],
            "Syllables per Word": [syllables_per_word],
            "Personal Pronouns Count": [personal_pronouns_count],
            "Average Word Length": [avg_word_length],
            "Positive score": [positive_score],
            "Negative score": [negative_score],
            "clened words": [clened_words]
        })

    else:
        # print("No. of Words:",np.nan)
        # print("Average Sentence Length:", np.nan)
        # print("Percentage of Complex Words:", np.nan)
        # print("FOG Index:", np.nan)
        # print("Average Number of Words per Sentence:", np.nan)
        # print("Complex Word Count:", np.nan)
        # print("Syllables per Word:", np.nan)
        # print("Personal Pronouns Count:", np.nan)
        # print("Average Word Length:", np.nan)
        # print("Positive score:", np.nan)
        # print("Negative score:", np.nan)
        # print("clened words:",np.nan)
        data =pd.DataFrame( {
             "URL":[url],
            "No. of Words": [np.nan],
            "Average Sentence Length": [np.nan],
            "Percentage of Complex Words": [np.nan],
            "FOG Index": [np.nan],
            "Average Number of Words per Sentence": [np.nan],
            "Complex Word Count": [np.nan],
            "Syllables per Word": [np.nan],
            "Personal Pronouns Count": [np.nan],
            "Average Word Length": [np.nan],
            "Positive score": [np.nan],
            "Negative score": [np.nan],
            "clened words": [np.nan]
        })
    return data


def count_syllables(word):
    vowels = 'aeiouy'
    count = 0
    prev_char = None
    for char in word:
        if char.lower() in vowels:
            if prev_char is None or prev_char.lower() not in vowels:
                count += 1
        prev_char = char
    if word.endswith(('es', 'ed')):
        count -= 1
    return count


# Test the function




# Load the Excel file
# Load the Excel file
workbook = openpyxl.load_workbook('input.xlsx')

# Select a specific worksheet
worksheet = workbook['Sheet1']

# Loop over rows with text or inputs
output = pd.DataFrame(columns=["URL", "Positive score", "Negative score", "Average Sentence Length",
                               "Percentage of Complex Words", "FOG Index", "Average Number of Words per Sentence",
                               "Complex Word Count", "Syllables per Word", "Personal Pronouns Count",
                               "Average Word Length", "clened words"])
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
    url = row[1]  # Assuming the URL is in the second column (index 1)
    if url is not None:
        print(url)
        x = calculate_score(url)
        output = output.append(x, ignore_index=True)
        print(output)
# Close the workbook
workbook.close()

# Print the output DataFrame
print(output)
output.to_csv("output.csv")
